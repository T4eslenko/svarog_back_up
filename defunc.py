import asyncio  
import os
import time
import openpyxl
from telethon.tl.functions.channels import InviteToChannelRequest
from telethon.tl.functions.contacts import GetContactsRequest, GetBlockedRequest
from telethon.tl.functions.messages import GetDialogsRequest, ImportChatInviteRequest
from telethon.tl.types import InputChannel, InputPhoneContact, User, Chat, Channel, Message, MessageFwdHeader, MessageMediaDocument, PeerChannel, DocumentAttributeFilename
from telethon.sync import TelegramClient, types
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from telethon.errors.rpcerrorlist import PeerFloodError, UserPrivacyRestrictedError
from datetime import datetime
from typing import Optional
import re
from jinja2 import Template
import base64
from io import BytesIO
from PIL import Image

async def get_user_info(client, phone, selection):
    """Функция для получения информации о пользователе и его ID."""
    me = await client.get_me()
    userid = me.id
    firstname = me.first_name
    username = f"@{me.username}" if me.username is not None else ""
    lastname = me.last_name if me.last_name is not None else ""
    userinfo = f"(Номер телефона: +{phone}, ID: {userid}, ({firstname} {lastname}) {username})"
    print("Информация о пользователе:") 
    print()
    print(f"Номер телефона: {phone}")
    print(f"ID пользователя: {userid}")
    print(f"Имя пользователя: {firstname} {lastname}")
    print(f"Username пользователя: {username}")
    if selection == '0':
        try:
            user_photo = await client.get_profile_photos(userid)
            if user_photo:
                file_name = f"{phone}.jpg"
                path = await client.download_media(user_photo[0], file=file_name)
        except Exception:
            pass
    return userid, userinfo, firstname, lastname, username


def get_blocked_bot(client, selection):
    blocked_bot_info = []
    blocked_bot_info_html = []
    count_blocked_bot = 0
    earliest_date = None
    latest_date = None
    image_data_url = " "
    
    delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, admin_id, user_bots, user_bots_html = get_type_of_chats(client, selection)
    result_blocked = client(GetBlockedRequest(offset=0, limit=200))
    for peer in result_blocked.blocked:
        if peer.peer_id.__class__.__name__ == 'PeerUser':
            user = client.get_entity(peer.peer_id.user_id)
            if user.bot:
                if selection == '0':
                    try:
                        photo_path = client.download_profile_photo(user, file=BytesIO())
                        if photo_path:
                            encoded_image = base64.b64encode(photo_path.getvalue()).decode('utf-8')
                            image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                        else:
                            img = Image.new('RGBA', (50, 50), (255, 255, 255, 0))
                            buffered = BytesIO()
                            img.save(buffered, format="PNG")
                            img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
                    except Exception:
                        pass    
                blocked_bot_info.append(f"\033[36m@{user.username}\033[0m \033[93m'{user.first_name}'\033[0m заблокирован: {peer.date.strftime('%d/%m/%Y')}")
                
                blocked_bot_info_html.append(
                    f'<img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
                    f'<a href="https://t.me/{user.username}" style="color:#0000FF; text-decoration: none;vertical-align:middle;">@{user.username}</a> '
                    f'<span style="color:#556B2F;vertical-align:middle;">{user.first_name}</span> заблокирован: {peer.date.strftime("%d/%m/%Y")}'
                )

    return count_blocked_bot, earliest_date, latest_date, blocked_bot_info, blocked_bot_info_html, user_bots, user_bots_html


def make_list_of_channels(delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, selection, client):
    """Функция для формирования списков групп и каналов"""
    owner_openchannel = 0
    owner_opengroup = 0
    owner_closegroup = 0
    owner_closechannel = 0
    all_info = []
    groups = []
    i=0


    openchannels_name = 'Открытые КАНАЛЫ:' if openchannels else ''
    all_info.append(f"\033[95m{openchannels_name}\033[0m")  
    openchannel_count = 1  
    public_channels_html = []
    image_data_url = ''
    for openchannel in openchannels:
        if selection == '0':
            try:
                photo_bytes = client.download_profile_photo(openchannel, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                        image_data_url = ''
            except Exception:
                pass 
        count_row = openchannel_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if openchannel.creator else ""
        admin = " (Администратор)" if openchannel.admin_rights is not None else ""
        messages_count = f" / [{chat_message_counts.get(openchannel.id, 0)}]" if chat_message_counts else ""
        all_info.append(f"{count_row} - {openchannel.title} \033[93m[{openchannel.participants_count}]{messages_count}\033[0m\033[91m {owner} {admin}\033[0m ID:{openchannel.id} \033[94m@{openchannel.username}\033[0m")
        public_channels_html.append(
        f"{openchannel_count}. <img src='{image_data_url}' alt=' ' style='width:50px;height:50px;vertical-align:middle;margin-right:10px;'>" 
        f"<span style='color:#556B2F;'>{openchannel.title}</span> <span style='color:#8B4513;'>[{openchannel.participants_count}]</span> "
        f"<span style='color:#FF0000;'>{owner} {admin}</span> ID:{openchannel.id} "
        f'<a href="https://t.me/{openchannel.username}" style="color:#0000FF; text-decoration: none;">@{openchannel.username}</a>'
        )
        openchannel_count += 1
        groups.append(openchannel)
        i +=1
        if owner != "" or admin != "":
            owner_openchannel += 1

    closechannels_name = 'Закрытые КАНАЛЫ:' if closechannels else ''
    all_info.append(f"\033[95m{closechannels_name}\033[0m")  
    closechannel_count = 1
    private_channels_html = []
    image_data_url = ''
    for closechannel in closechannels:
        if selection == '0':
            try:
                photo_bytes = client.download_profile_photo(closechannel, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                    img = Image.new('RGBA', (50, 50), (255, 255, 255, 0))
                    buffered = BytesIO()
                    img.save(buffered, format="PNG")
                    img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
                    image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = closechannel_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if closechannel.creator else ""
        admin = " (Администратор)" if closechannel.admin_rights is not None else ""
        messages_count = f" / [{chat_message_counts.get(closechannel.id, 0)}]" if chat_message_counts else ""
        all_info.append(f"{count_row} - {closechannel.title} \033[93m[{closechannel.participants_count}]{messages_count}\033[0m \033[91m{owner} {admin}\033[0m ID:{closechannel.id}")
        private_channels_html.append(
            f'{closechannel_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
            f"<span style='color:#556B2F;'>{closechannel.title}</span> <span style='color:#8B4513;'>[{closechannel.participants_count}]</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{closechannel.id}"
        )
        closechannel_count += 1
        groups.append(closechannel)
        i +=1
        if owner != "" or admin != "":
            owner_closechannel += 1

    openchats_name = 'Открытые ГРУППЫ:' if openchats else ''
    all_info.append(f"\033[95m{openchats_name}\033[0m")
    opengroup_count = 1
    public_groups_html = []
    image_data_url = ''
    for openchat in openchats:
        if selection == '0':
            try:
                photo_bytes = client.download_profile_photo(openchat, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                    img = Image.new('RGBA', (50, 50), (255, 255, 255, 0))
                    buffered = BytesIO()
                    img.save(buffered, format="PNG")
                    img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
                    image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = opengroup_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if openchat.creator else ""
        admin = " (Администратор)" if openchat.admin_rights is not None else ""
        messages_count = f" / [{chat_message_counts.get(openchat.id, 0)}]" if chat_message_counts else ""
        all_info.append(f"{count_row} - {openchat.title} \033[93m[{openchat.participants_count}]{messages_count}\033[0m\033[91m {owner} {admin}\033[0m ID:{openchat.id} \033[94m@{openchat.username}\033[0m")
        public_groups_html.append(
            f'{opengroup_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
            f"<span style='color:#556B2F;'>{openchat.title}</span> <span style='color:#8B4513;'>[{openchat.participants_count}]</span> "
            f"<span style='color:#FF0000;'>{owner} {admin}</span> ID:{openchat.id} "
            f'<a href="https://t.me/{openchat.username}" style="color:#0000FF; text-decoration: none;">@{openchat.username}</a>'
        )
        opengroup_count += 1
        groups.append(openchat)
        i +=1
        if owner != "" or admin != "":
            owner_opengroup += 1

    closechats_name = 'Закрытые ГРУППЫ:' if closechats else ''
    all_info.append(f"\033[95m{closechats_name}\033[0m")
    closegroup_count = 1
    private_groups_html = []
    image_data_url = ''
    for closechat in closechats:
        if selection == '0':
            try:
                photo_bytes = client.download_profile_photo(closechat, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                    img = Image.new('RGBA', (50, 50), (255, 255, 255, 0))
                    buffered = BytesIO()
                    img.save(buffered, format="PNG")
                    img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
                    image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = closegroup_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if closechat.creator else ""
        admin = " (Администратор)" if closechat.admin_rights is not None else ""
        messages_count = f" / [{chat_message_counts.get(closechat.id, 0)}]" if chat_message_counts else ""
        all_info.append(f"{count_row} - {closechat.title} \033[93m[{closechat.participants_count}]{messages_count}\033[0m \033[91m{owner} {admin}\033[0m ID:{closechat.id}")
        #private_groups_html.append(f"{closegroup_count} - <span style='color:#556B2F;'>{closechat.title}</span> <span style='color:#8B4513;'>[{closechat.participants_count}]</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{closechat.id}")
        private_groups_html.append(
            f'{closegroup_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
            f"<span style='color:#556B2F;'>{closechat.title}</span> <span style='color:#8B4513;'>[{closechat.participants_count}]</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{closechat.id}"
        )
        closegroup_count += 1
        groups.append(closechat)
        i +=1
        if owner != "" or admin != "":
            owner_closegroup += 1

    
    delgroups_name = 'Удаленные ГРУППЫ:' if delgroups else ''
    all_info.append(f"\033[95m{delgroups_name}\033[0m")
    closegroupdel_count = 1
    deleted_groups_html = []
    for delgroup in delgroups:
        count_row = closegroupdel_count if selection == '5' or selection == '0' else i
        owner_value = delgroup['creator']
        admin_value = delgroup['admin_rights']
        id_value = delgroup['ID']
        title_value = delgroup['title']
        owner = " (Владелец)" if owner_value else ""
        admin = " (Администратор)" if admin_value is not None else ""
        all_info.append(f"{count_row} - {title_value} \033[91m{owner} {admin}\033[0m ID:{id_value}")
        deleted_groups_html.append(f"{closegroupdel_count} - <span style='color:#556B2F;'>{title_value}</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{id_value}")
        closegroupdel_count += 1
        i +=1
        if owner != "" or admin != "":
            owner_closegroup += 1

    return groups, i, all_info, openchannel_count, closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, public_channels_html, private_channels_html, public_groups_html, private_groups_html, deleted_groups_html

def get_and_save_contacts(client, phone, userinfo, userid):
    result = client(GetContactsRequest(0))
    contacts = result.users
    total_contacts = len(contacts)
    total_contacts_with_phone = sum(bool(getattr(contact, 'phone', None)) for contact in contacts)
    total_mutual_contacts = sum(bool(getattr(contact, 'mutual_contact', None)) for contact in contacts)
    print(f"Количество контактов: {total_contacts}")
    print(f"Количество контактов с номерами телефонов: {total_contacts_with_phone}")
    print(f"Количество взаимных контактов: {total_mutual_contacts}")
    print()
    
    # Сохраняем информацию о контактах
    contacts_file_name = f'{phone}_contacts.xlsx'
    print(f"Контакты сохранены в файл {phone}_contacts.xlsx")

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value=userinfo)
    headers = ['ID', 'First name (так записан у объекта в книге)', 'Last name (так записан у объекта в книге)', 'Username', 'Телефон', 'Взаимный контакт', 'Дата внесения в базу', 'ID объекта']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col, value=header)
        
    row_num = 3
    for contact in contacts:
        if hasattr(contact, 'id'):
            sheet.cell(row=row_num, column=1, value=contact.id)
        if hasattr(contact, 'first_name'):
            sheet.cell(row=row_num, column=2, value=contact.first_name)
        if hasattr(contact, 'last_name'):
            sheet.cell(row=row_num, column=3, value=contact.last_name)
        if hasattr(contact, 'username') and contact.username is not None:
            username_with_at = f"@{contact.username}"
            sheet.cell(row=row_num, column=4, value=username_with_at)
        if hasattr(contact, 'phone'):
            sheet.cell(row=row_num, column=5, value=contact.phone)
        if hasattr(contact, 'mutual_contact') and contact.mutual_contact:
            sheet.cell(row=row_num, column=6, value='взаимный')
        sheet.cell(row=row_num, column=7, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row_num, column=8, value=userid)
     
        row_num += 1

    wb.save(contacts_file_name)
    return total_contacts, total_contacts_with_phone, total_mutual_contacts

def save_about_channels(phone, userid, firstname, lastname, username, openchannel_count, opengroup_count, closechannel_count, closegroup_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, openchannels, closechannels, openchats, closechats, delgroups, closegroupdel_count):
    
    def write_data(sheet, data):
        sheet.append(["Название", "Количество участников", "Владелец", "Администратор", "ID", "Ссылка"])
        for item in data:
          owner = " (Владелец)" if item.creator else ""
          admin = " (Администратор)" if item.admin_rights is not None else ""
          usernameadd = f"@{item.username}" if hasattr(item, 'username') and item.username is not None else ""
          sheet.append([item.title, item.participants_count, owner, admin, item.id, usernameadd])
    
    def write_data_del(sheet, data):
        sheet.append(["Название", "Владелец", "Администратор", "ID"])
        for item in data:
          owner_value = item['creator']
          admin_value = item['admin_rights']
          id_value = item['ID']
          title_value = item['title']
          owner = " (Владелец)" if owner_value else ""
          admin = " (Администратор)" if admin_value is not None else ""
          sheet.append([title_value, owner, admin, id_value])
            
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_summury = wb.create_sheet("Сводная информация")
    ws_summury.append([f"Номер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}"])
    if openchannel_count > 1:
        ws_summury.append([f"Открытые каналы: {openchannel_count-1}"])
        ws_open_channels = wb.create_sheet("Открытые каналы")
        write_data(ws_open_channels, openchannels)
    if closechannel_count > 1:
        ws_summury.append([f"Закрытые каналы: {closechannel_count-1}"])
        ws_closed_channels = wb.create_sheet("Закрытые каналы")
        write_data(ws_closed_channels, closechannels)
    if owner_openchannel > 1:
        ws_summury.append([f"Имеет права владельца или админа в открытых каналах: {owner_openchannel}"])
    if owner_closechannel > 1:
        ws_summury.append([f"Имеет права владельца или админа в закрытых каналах: {owner_closechannel}"])
    if opengroup_count > 1:
        ws_summury.append([f"Открытые группы: {opengroup_count-1}"])
        ws_open_groups = wb.create_sheet("Открытые группы")
        write_data(ws_open_groups, openchats)
    if closegroup_count > 1:
        ws_summury.append([f"Закрытые группы: {closegroup_count-1}"])
        ws_closed_groups = wb.create_sheet("Закрытые группы")
        write_data(ws_closed_groups, closechats)
    if closegroupdel_count > 1:
        ws_summury.append([f"Удаленные группы: {closegroupdel_count-1}"])
        ws_closed_groups_del = wb.create_sheet("Удаленные группы")
        write_data_del(ws_closed_groups_del, delgroups)
    if owner_opengroup > 11:
        ws_summury.append([f"Имеет права владельца или админа в открытых группах: {owner_opengroup}"])
    if owner_closegroup > 1:
        ws_summury.append([f"Имеет права владельца или админа в закрытых группах: {owner_closegroup}"])
    
    wb.save(f"{phone}_about.xlsx")

#  Формируем отчет HTML
def generate_html_report(phone, userid, userinfo, firstname, lastname, username, total_contacts, total_contacts_with_phone, total_mutual_contacts, openchannel_count, closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, public_channels_html, private_channels_html, public_groups_html, private_groups_html, deleted_groups_html, blocked_bot_info_html, user_bots_html):
    # Путь к аватарке пользователя
    avatar_path = f"{phone}.jpg"
    
    if os.path.exists(avatar_path):
        # Чтение и конвертация изображения в Base64
        with open(avatar_path, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            avatar_data_uri = f"data:image/jpeg;base64,{encoded_string}"
    else:
        # Используем изображение по умолчанию или оставляем поле пустым
        avatar_data_uri = "data:image/gif;base64,R0lGODlhAQABAIAAAAUEBA=="  # 1x1 прозрачный GIF
    
    # Открываем HTML шаблон
    with open('template.html', 'r', encoding='utf-8') as file:
        template = Template(file.read())

    # Заполняем шаблон данными
    html_content = template.render(
        phone=phone,
        userid=userid,
        firstname=firstname,
        lastname=lastname,
        username=username,
        total_contacts=total_contacts,
        total_contacts_with_phone=total_contacts_with_phone,
        total_mutual_contacts=total_mutual_contacts,
        openchannel_count=openchannel_count,
        closechannel_count=closechannel_count,
        opengroup_count=opengroup_count,
        closegroup_count=closegroup_count,
        closegroupdel_count=closegroupdel_count,
        owner_openchannel=owner_openchannel,
        owner_closechannel=owner_closechannel,
        owner_opengroup=owner_opengroup,
        owner_closegroup=owner_closegroup,
        blocked_bot_info_html=blocked_bot_info_html,
        user_bots_html=user_bots_html,
        public_channels_html=public_channels_html,
        private_channels_html=private_channels_html,
        public_groups_html=public_groups_html,
        private_groups_html=private_groups_html,
        deleted_groups_html=deleted_groups_html,
        avatar_user=avatar_data_uri
    )

    # Сохраняем результат в HTML файл
    report_filename = f"{phone}_report.html"
    with open(report_filename, 'w', encoding='utf-8') as file:
        file.write(html_content)    
